"""
IMSUCC Result Analysis System
Flexible for any course, any semester, any number of subjects.
Input: Standard Result Format Excel (Sheet 1 = student data, Sheet 2 = subject-faculty map)
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="IMSUCC – Result Analysis System",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
#  GLOBAL STYLES
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html,body,[class*="css"]{font-family:'Inter',sans-serif}

/* Institute banner */
.inst-banner{
  background:linear-gradient(135deg,#0d1b5e 0%,#1a237e 55%,#283593 100%);
  color:#fff;padding:1.4rem 2rem 1rem;border-radius:16px;
  text-align:center;margin-bottom:.6rem;
  box-shadow:0 6px 28px rgba(13,27,94,.35)
}
.inst-banner .inst-name{font-size:1.35rem;font-weight:700;letter-spacing:.3px;margin-bottom:.15rem}
.inst-banner .inst-city{font-size:.85rem;opacity:.82;letter-spacing:.6px;text-transform:uppercase}

/* Report title bar */
.report-bar{
  background:#1a237e;color:#fff;padding:.75rem 2rem;
  border-radius:10px;text-align:center;margin-bottom:1.2rem;font-size:1.05rem;font-weight:600
}

/* KPI card */
.kpi{background:#fff;border-radius:12px;padding:1rem 1.2rem;
  box-shadow:0 2px 10px rgba(0,0,0,.09);border-left:5px solid;margin-bottom:.5rem}
.kpi-val{font-size:1.85rem;font-weight:700}
.kpi-lbl{font-size:.72rem;color:#666;text-transform:uppercase;letter-spacing:.5px;margin-top:.1rem}

/* Section header */
.sec{font-size:1.05rem;font-weight:600;color:#1a237e;
  border-bottom:2px solid #e8eaf6;padding-bottom:.35rem;margin:1rem 0 .7rem}

/* Info box */
.info-box{background:#e8eaf6;border-radius:8px;padding:.75rem 1.1rem;
  border-left:4px solid #3949ab;margin:.7rem 0;font-size:.88rem}

/* Upload area */
.upload-area{border:2px dashed #3949ab;border-radius:12px;padding:2.2rem;
  text-align:center;background:#f3f4ff;margin:1rem 0}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
INSTITUTE_NAME   = "Institute of Management Studies University Courses Campus (IMSUCC)"
INSTITUTE_CITY   = "Ghaziabad"
COLORS           = ['#1a237e','#283593','#3949ab','#5c6bc0','#7986cb','#9fa8da','#c5cae9']

# Ordinal helper
def ordinal(n):
    n = int(n)
    return f"{n}{'th' if 11<=n%100<=13 else {1:'st',2:'nd',3:'rd'}.get(n%10,'th')}"

# ─────────────────────────────────────────────────────────────────────────────
#  PARSING — fully dynamic, no hard-coded subject codes
# ─────────────────────────────────────────────────────────────────────────────

def flatten_multiindex(df):
    """Flatten a two-level column header into unique, readable names."""
    flat, seen = [], {}
    for top, sub in df.columns:
        t = str(top).strip().replace('\n', ' ')
        s = str(sub).strip().replace('\n', ' ')
        if 'Unnamed' in t: t = ''
        if 'Unnamed' in s: s = ''
        key = f"{t}_{s}".strip('_') or '_'
        seen[key] = seen.get(key, 0) + 1
        flat.append(key if seen[key] == 1 else f"{key}_{seen[key]}")
    df.columns = flat
    return df


def detect_subject_ttl_cols(df):
    """
    Auto-detect subject 'Ttl' columns from the flattened DataFrame.
    Pattern: column name like '<code>_Ttl' where code is numeric (e.g. 501_Ttl).
    Returns dict: {code_str: col_name}
    """
    pattern = re.compile(r'^(\d{3,4})_Ttl$', re.IGNORECASE)
    result = {}
    for col in df.columns:
        m = pattern.match(col)
        if m:
            result[m.group(1)] = col
    return result


def detect_special_cols(df):
    """
    Find practical/project columns (505, 506, 507, 508 style).
    Returns dict: {code_str: col_name}
    """
    pattern = re.compile(r'^(\d{3,4})_\(P\)|^(\d{3,4})_Minor|^(\d{3,4})_Summar', re.IGNORECASE)
    result = {}
    for col in df.columns:
        m = pattern.match(col)
        if m:
            code = next(g for g in m.groups() if g)
            result[code] = col
    return result


def find_col(df, keywords, exact=False):
    """Return first column name containing any of the keywords (case-insensitive)."""
    for col in df.columns:
        c = col.lower()
        for kw in keywords:
            if (c == kw.lower()) if exact else (kw.lower() in c):
                return col
    return None


def parse_sheet1(uploaded_file):
    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, sheet_name=0, header=[0, 1])
    df = flatten_multiindex(df)
    sno_col = find_col(df, ['s. no','sno','s.no'])
    if sno_col:
        df = df.dropna(subset=[sno_col])
    df = df.reset_index(drop=True)

    # Convert all numeric-looking columns
    for col in df.columns:
        try:
            df[col] = pd.to_numeric(df[col], errors='ignore')
        except Exception:
            pass

    # Force-convert subject Ttl cols + total/percent cols
    subj_cols = detect_subject_ttl_cols(df)
    for col in subj_cols.values():
        df[col] = pd.to_numeric(df[col], errors='coerce')

    for kw in ['Totel', 'Grand Total', 'PER']:
        col = find_col(df, [kw])
        if col:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    return df


def parse_sheet2(uploaded_file):
    uploaded_file.seek(0)
    raw = pd.read_excel(uploaded_file, sheet_name=1, header=None)
    df = raw.iloc[1:].copy()
    # Sheet 2 always: Subject Code | Subject Name | Faculty Name | Sections
    df.columns = ['Subject Code', 'Subject Name', 'Faculty Name', 'Sections']
    df = df.dropna(subset=['Subject Code'])
    df['Subject Code'] = df['Subject Code'].astype(str).str.strip()
    df['Subject Name'] = df['Subject Name'].astype(str).str.strip()
    df['Faculty Name'] = df['Faculty Name'].astype(str).str.strip()
    df['Sections']     = df['Sections'].astype(str).str.strip()
    return df.reset_index(drop=True)

# ─────────────────────────────────────────────────────────────────────────────
#  ANALYSIS — fully dynamic
# ─────────────────────────────────────────────────────────────────────────────

def result_summary(df, sem_label):
    status_col = find_col(df, ['status'])
    if not status_col:
        return {}
    s       = df[status_col].astype(str).str.strip().str.upper()
    total   = len(df)
    passed  = int((s == 'PASS').sum())
    later   = int((s == 'LATER').sum())
    bp      = int((s == 'BP').sum())
    fail    = int((s == 'FAIL').sum())
    ufm     = int((s == 'UFM').sum())
    pct_v   = round(passed / total * 100, 2) if total else 0
    pct_all = round((passed + later) / total * 100, 2) if total else 0
    return {
        'Total Registered': total,
        'Total Appeared':   total,
        f'Pass ({sem_label})': passed,
        'Later / Overall Pass': later,
        'Back Paper':       bp,
        'Fail':             fail,
        'UFM':              ufm,
        f'% Pass in {sem_label}': pct_v,
        '% Pass Overall (All Sems)': pct_all,
    }


def compute_subject_stats(df, sf_df):
    subj_cols = detect_subject_ttl_cols(df)
    records   = []
    for code, col in subj_cols.items():
        vals     = pd.to_numeric(df[col], errors='coerce').dropna()
        appeared = len(vals)
        if not appeared:
            continue
        back     = int((vals < 40).sum())
        passed   = appeared - back
        pct_pass = round(passed / appeared * 100, 2)
        avg      = round(float(vals.mean()), 2)
        above60  = round((vals >= 60).sum() / appeared * 100, 2)
        highest  = int(vals.max())

        fac_rows  = sf_df[sf_df['Subject Code'].astype(str).str.contains(code, regex=False)]
        subj_name = fac_rows['Subject Name'].dropna().iloc[0] if not fac_rows.empty else code
        faculties = ', '.join(fac_rows['Faculty Name'].dropna().tolist()) if not fac_rows.empty else 'N/A'

        records.append({
            'Code': code,
            'Subject Name': str(subj_name).strip()[:40],
            'Faculty': faculties,
            'Appeared': appeared,
            'Back Paper': back,
            'Pass': passed,
            'Pass %': pct_pass,
            'Avg Score': avg,
            '% > 60': above60,
            'Highest Marks': highest,
        })
    return pd.DataFrame(records)


def compute_faculty_stats(df, sf_df):
    subj_cols = detect_subject_ttl_cols(df)
    grp_col   = find_col(df, ['GR.', 'gr.', 'group', 'section'], exact=False)
    records   = []
    for _, row in sf_df.iterrows():
        code     = str(row['Subject Code']).strip()
        faculty  = str(row['Faculty Name']).strip()
        sections = str(row['Sections']).strip()
        subj     = str(row['Subject Name']).strip()
        col      = subj_cols.get(code)
        if not col or col not in df.columns:
            continue
        if grp_col and sections not in ('nan', '', 'NaN'):
            sec_list = [s.strip() for s in sections.split(',')]
            sub_df   = df[df[grp_col].astype(str).str.strip().isin(sec_list)]
        else:
            sub_df = df
        vals     = pd.to_numeric(sub_df[col], errors='coerce').dropna()
        appeared = len(vals)
        if not appeared:
            continue
        back   = int((vals < 40).sum())
        passed = appeared - back
        pct    = round(passed / appeared * 100, 2)
        records.append({
            'S.No':         len(records) + 1,
            'Faculty Name': faculty,
            'Sections':     sections,
            'Subject':      subj[:35],
            'Code':         code,
            'Appeared':     appeared,
            'Back Paper':   back,
            'Pass':         passed,
            'Pass %':       pct,
        })
    return pd.DataFrame(records)


def compute_grade_dist(df, sf_df):
    subj_cols = detect_subject_ttl_cols(df)
    records   = []
    for code, col in subj_cols.items():
        vals     = pd.to_numeric(df[col], errors='coerce').dropna()
        appeared = len(vals)
        if not appeared:
            continue
        fac_rows = sf_df[sf_df['Subject Code'].astype(str).str.contains(code, regex=False)]
        subj     = fac_rows['Subject Name'].dropna().iloc[0] if not fac_rows.empty else code
        lt50  = int((vals < 50).sum())
        b5060 = int(((vals >= 50) & (vals < 60)).sum())
        b6075 = int(((vals >= 60) & (vals < 75)).sum())
        ab75  = int((vals >= 75).sum())
        records.append({
            'Subject':     f"{code} – {str(subj).strip()[:22]}",
            '< 50':        lt50,
            '50–59.9':     b5060,
            '60–74.9':     b6075,
            '≥ 75':        ab75,
            '< 50 %':      round(lt50  / appeared * 100, 1),
            '50–59.9 %':   round(b5060 / appeared * 100, 1),
            '60–74.9 %':   round(b6075 / appeared * 100, 1),
            '≥ 75 %':      round(ab75  / appeared * 100, 1),
        })
    return pd.DataFrame(records)


def compute_toppers(df, n=10):
    grand_col = find_col(df, ['grand total', 'grand_total'])
    sem_col   = find_col(df, ['totel', 'total_v', 'totel (v)'])
    sort_col  = grand_col or sem_col
    if not sort_col:
        return pd.DataFrame()
    name_col = find_col(df, ['candidate name', 'candidate_name', 'name'])
    roll_col = find_col(df, ['roll no', 'roll_no', 'rollno'])
    pct_col  = find_col(df, ['per_%.1', 'per_%', 'percentage', 'per'])

    sel = [c for c in [roll_col, name_col, sort_col, pct_col] if c]
    out = df[sel].copy()
    out[sort_col] = pd.to_numeric(out[sort_col], errors='coerce')
    top = out.nlargest(n, sort_col).reset_index(drop=True)
    top.index += 1; top.index.name = 'Rank'
    return top.reset_index()

# ─────────────────────────────────────────────────────────────────────────────
#  CHARTS
# ─────────────────────────────────────────────────────────────────────────────

def _layout(fig, title, h=400):
    fig.update_layout(
        title=title, title_x=.5,
        plot_bgcolor='white', paper_bgcolor='white',
        font=dict(family='Inter', size=12),
        height=h, margin=dict(t=65, b=55, l=50, r=20),
    )
    return fig


def chart_pie(summary):
    cats   = [k for k in summary if k not in ('Total Registered','Total Appeared') and '%' not in k]
    vals   = [summary[k] for k in cats]
    colors = ['#2e7d32','#1565c0','#e65100','#c62828','#6a1b9a','#37474f']
    fig = go.Figure(go.Pie(
        labels=cats, values=vals, hole=.42,
        marker=dict(colors=colors[:len(cats)]),
        hovertemplate='%{label}: %{value} (%{percent})<extra></extra>',
    ))
    return _layout(fig, 'Result Status Breakdown', 360)


def chart_bar_summary(summary):
    cats   = [k for k in summary if k not in ('Total Registered','Total Appeared') and '%' not in k]
    vals   = [summary[k] for k in cats]
    colors = ['#2e7d32','#1565c0','#e65100','#c62828','#6a1b9a','#37474f']
    fig = go.Figure(go.Bar(
        x=cats, y=vals, marker_color=colors[:len(cats)],
        text=vals, textposition='outside',
    ))
    fig.update_yaxes(title='No. of Students')
    return _layout(fig, 'Overall Result Distribution', 360)


def chart_subject_pass(sdf):
    fig = go.Figure()
    fig.add_trace(go.Bar(name='Pass %', x=sdf['Code'], y=sdf['Pass %'],
                         marker_color='#1a237e', text=sdf['Pass %'], textposition='outside'))
    fig.add_trace(go.Bar(name='% > 60', x=sdf['Code'], y=sdf['% > 60'],
                         marker_color='#7986cb', text=sdf['% > 60'], textposition='outside'))
    fig.update_layout(barmode='group', yaxis_range=[0, 118], yaxis_title='Percentage')
    return _layout(fig, 'Subject Wise: Pass % vs Students Scoring Above 60%', 400)


def chart_avg(sdf):
    fig = go.Figure(go.Bar(
        x=sdf['Code'], y=sdf['Avg Score'],
        marker_color=COLORS[:len(sdf)],
        text=sdf['Avg Score'], textposition='outside',
    ))
    fig.update_yaxes(title='Average Score (out of 100)')
    return _layout(fig, 'Average Score per Subject', 380)


def chart_grade_dist(gdf):
    bands  = ['< 50 %', '50–59.9 %', '60–74.9 %', '≥ 75 %']
    bcolor = ['#c62828', '#e65100', '#1565c0', '#2e7d32']
    fig = go.Figure()
    for band, col in zip(bands, bcolor):
        fig.add_trace(go.Bar(name=band, x=gdf['Subject'], y=gdf[band], marker_color=col))
    fig.update_layout(
        barmode='stack', yaxis_title='% of Students',
        legend=dict(orientation='h', y=-0.32),
        margin=dict(t=65, b=130, l=50, r=20), height=460,
    )
    return _layout(fig, 'Score Band Distribution by Subject (%)', 460)


def chart_faculty(fdf):
    fig = go.Figure(go.Bar(
        x=fdf['Faculty Name'], y=fdf['Pass %'],
        marker_color=[COLORS[i % len(COLORS)] for i in range(len(fdf))],
        text=fdf['Pass %'], textposition='outside',
    ))
    fig.update_layout(
        xaxis_tickangle=-32, yaxis_range=[0, 118], yaxis_title='Pass %',
        margin=dict(t=65, b=140, l=50, r=20), height=460,
    )
    return _layout(fig, 'Faculty Wise Pass %', 460)


def chart_score_hist(df):
    pct_col = find_col(df, ['per_%', 'percentage', 'per'])
    if not pct_col:
        return None
    vals = pd.to_numeric(df[pct_col], errors='coerce').dropna()
    fig = go.Figure(go.Histogram(x=vals, nbinsx=20, marker_color='#3949ab'))
    fig.update_xaxes(title='Semester Percentage')
    fig.update_yaxes(title='No. of Students')
    return _layout(fig, 'Distribution of Student Scores (%)', 370)


def chart_toppers(top_df):
    name_col  = next((c for c in top_df.columns if 'name' in c.lower()), None)
    score_col = next((c for c in top_df.columns if 'grand' in c.lower() or 'total' in c.lower()
                      or 'totel' in c.lower()), None)
    if not name_col or not score_col:
        return None
    fig = go.Figure(go.Bar(
        x=top_df[name_col].astype(str),
        y=pd.to_numeric(top_df[score_col], errors='coerce'),
        marker_color=COLORS[:len(top_df)],
        text=top_df[score_col], textposition='outside',
    ))
    fig.update_layout(
        xaxis_tickangle=-28, yaxis_title='Grand Total',
        margin=dict(t=65, b=130, l=50, r=20), height=440,
    )
    return _layout(fig, 'Top Students by Grand Total', 440)

# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(course, sem_label, summary, sdf, fdf, gdf, top_df, student_df):
    wb = openpyxl.Workbook()
    HF = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    HB = PatternFill('solid', start_color='1A237E')
    TF = Font(bold=True, name='Arial', size=13, color='1A237E')
    SF = Font(bold=True, name='Arial', size=11, color='1A237E')
    DF = Font(name='Arial', size=10)
    CTR= Alignment(horizontal='center', vertical='center', wrap_text=True)
    LFT= Alignment(horizontal='left',   vertical='center', wrap_text=True)
    TH = Side(style='thin', color='CCCCCC')
    BD = Border(left=TH, right=TH, top=TH, bottom=TH)
    ALT= PatternFill('solid', start_color='F5F5F5')

    def hdr(ws, row, cols):
        for j, v in enumerate(cols, 1):
            c = ws.cell(row=row, column=j, value=v)
            c.font=HF; c.fill=HB; c.alignment=CTR; c.border=BD

    def drow(ws, row, vals, alt=False):
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=j, value=v)
            c.font=DF; c.alignment=CTR; c.border=BD
            if alt: c.fill=ALT

    def aw(ws):
        for col in ws.columns:
            ml = max((len(str(cell.value or '')) for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, 42)

    def titl(ws, text, sub, ncols):
        cl = get_column_letter(max(ncols, 1))
        ws.merge_cells(f'A1:{cl}1')
        c = ws['A1']; c.value=text; c.font=TF; c.alignment=CTR
        ws.merge_cells(f'A2:{cl}2')
        d = ws['A2']; d.value=sub; d.font=Font(italic=True,name='Arial',size=9,color='444444')
        d.alignment=CTR

    sub_hdr = f"{INSTITUTE_NAME}, {INSTITUTE_CITY}  |  {course}  |  {sem_label} Semester Result Analysis"

    # Sheet 1: Summary
    ws1 = wb.active; ws1.title = 'Overall Summary'
    titl(ws1, 'RESULT ANALYSIS – OVERALL SUMMARY', sub_hdr, 3)
    hdr(ws1, 4, ['Metric', 'Value'])
    for i, (k, v) in enumerate(summary.items(), 5):
        ws1.cell(row=i, column=1, value=k).font = Font(bold=True, name='Arial', size=10)
        ws1.cell(row=i, column=1).border = BD
        ws1.cell(row=i, column=1).alignment = LFT
        c = ws1.cell(row=i, column=2, value=v)
        c.font = Font(bold=True, name='Arial', size=11, color='1A237E')
        c.alignment = CTR; c.border = BD
        if i % 2 == 0: ws1.cell(row=i, column=1).fill = ALT; c.fill = ALT
    ws1.column_dimensions['A'].width = 32; ws1.column_dimensions['B'].width = 18

    # Sheet 2: Subject
    ws2 = wb.create_sheet('Subject Wise Analysis')
    if not sdf.empty:
        titl(ws2, 'SUBJECT WISE RESULT ANALYSIS', sub_hdr, len(sdf.columns))
        hdr(ws2, 4, list(sdf.columns))
        for i, row in sdf.iterrows(): drow(ws2, i+5, list(row), i%2==0)
    aw(ws2)

    # Sheet 3: Faculty
    ws3 = wb.create_sheet('Faculty Wise Analysis')
    if not fdf.empty:
        titl(ws3, 'FACULTY WISE RESULT ANALYSIS', sub_hdr, len(fdf.columns))
        hdr(ws3, 4, list(fdf.columns))
        for i, row in fdf.iterrows(): drow(ws3, i+5, list(row), i%2==0)
    aw(ws3)

    # Sheet 4: Grade dist
    ws4 = wb.create_sheet('Grade Distribution')
    if not gdf.empty:
        dc = ['Subject','< 50','50–59.9','60–74.9','≥ 75','< 50 %','50–59.9 %','60–74.9 %','≥ 75 %']
        titl(ws4, 'SCORE BAND DISTRIBUTION', sub_hdr, len(dc))
        hdr(ws4, 4, dc)
        for i, row in gdf[dc].iterrows(): drow(ws4, i+5, list(row), i%2==0)
    aw(ws4)

    # Sheet 5: Toppers
    ws5 = wb.create_sheet('Toppers')
    if not top_df.empty:
        titl(ws5, 'TOP STUDENTS', sub_hdr, len(top_df.columns))
        hdr(ws5, 4, list(top_df.columns))
        for i, row in top_df.iterrows(): drow(ws5, i+5, list(row), i%2==0)
    aw(ws5)

    # Sheet 6: Full data
    ws6 = wb.create_sheet('Student Data')
    titl(ws6, 'COMPLETE STUDENT RESULT DATA', sub_hdr, len(student_df.columns))
    hdr(ws6, 4, list(student_df.columns))
    for i, row in student_df.iterrows():
        for j, val in enumerate(row, 1):
            c = ws6.cell(row=i+5, column=j, value=val)
            c.font=DF; c.alignment=CTR; c.border=BD
            if i % 2 == 0: c.fill=ALT
    aw(ws6)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
#  PRINT-READY HTML
# ─────────────────────────────────────────────────────────────────────────────

def build_print_html(course, sem_label, summary, sdf, fdf, gdf, top_df):
    def tbl(df, title, sel=None):
        if df is None or df.empty:
            return f'<h3 class="stl">{title}</h3><p>No data available.</p>'
        d = df[sel] if sel else df
        th   = ''.join(f'<th>{c}</th>' for c in d.columns)
        rows = ''.join(
            '<tr>' + ''.join(f'<td>{v}</td>' for v in r) + '</tr>'
            for _, r in d.iterrows()
        )
        return f'<h3 class="stl">{title}</h3><table><thead><tr>{th}</tr></thead><tbody>{rows}</tbody></table>'

    sum_rows = ''.join(
        f'<tr><td class="lb">{k}</td><td class="vl">{v}</td></tr>'
        for k, v in summary.items()
    )
    gc = ['Subject', '< 50 %', '50–59.9 %', '60–74.9 %', '≥ 75 %']

    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<title>Result Analysis Report – {course} {sem_label} Sem</title>
<style>
@page{{size:A4;margin:13mm}}
body{{font-family:Arial,sans-serif;font-size:9.5pt;color:#000;background:#fff;margin:0}}
/* Institute header */
.inst-hdr{{text-align:center;margin-bottom:10px}}
.inst-hdr h1{{font-size:13pt;font-weight:bold;margin:0 0 2px;letter-spacing:.3px}}
.inst-hdr h2{{font-size:11pt;font-weight:normal;margin:0 0 2px}}
.inst-hdr .city{{font-size:9pt;color:#333;margin:0}}
.divider{{border:none;border-top:3px solid #000;margin:6px 0 10px}}
/* Section titles */
.stl{{font-size:11pt;font-weight:bold;margin:14px 0 5px;
      border-bottom:2px solid #333;padding-bottom:3px;page-break-after:avoid}}
/* Tables */
table{{width:100%;border-collapse:collapse;margin-bottom:13px;font-size:8.5pt}}
th{{background:#333;color:#fff;padding:4px 6px;border:1px solid #444;text-align:center}}
td{{padding:3px 5px;border:1px solid #bbb;text-align:center}}
tr:nth-child(even) td{{background:#f5f5f5}}
.lb{{font-weight:bold;background:#ebebeb;text-align:left;padding-left:8px}}
.vl{{font-weight:bold}}
/* Footer */
.ft{{text-align:center;margin-top:16px;font-size:7.5pt;color:#555;
    border-top:1px solid #bbb;padding-top:5px}}
.no-print{{margin-bottom:12px}}
@media print{{.no-print{{display:none}}}}
</style></head>
<body>

<div class="inst-hdr">
  <h1>{INSTITUTE_NAME}</h1>
  <h2 class="city">{INSTITUTE_CITY}</h2>
  <hr class="divider">
  <h2>Result Analysis &nbsp;|&nbsp; {course} &nbsp;|&nbsp; {sem_label} Semester</h2>
</div>

<div class="no-print" style="margin-bottom:10px">
  <button onclick="window.print()"
    style="padding:7px 20px;font-size:11pt;cursor:pointer;
           background:#1a237e;color:#fff;border:none;border-radius:6px">
    🖨 Print / Save as PDF
  </button>
</div>

<h3 class="stl">Overall Result Summary</h3>
<table style="width:50%"><tbody>{sum_rows}</tbody></table>

{tbl(sdf, 'Subject Wise Result Analysis')}
{tbl(fdf, 'Faculty Wise Result Analysis')}
{tbl(gdf, 'Grade Distribution (%)', gc)}
{tbl(top_df, 'Top Students')}

<div class="ft">
  {INSTITUTE_NAME}, {INSTITUTE_CITY} &nbsp;|&nbsp; Result Analysis System
</div>
</body></html>"""

# ─────────────────────────────────────────────────────────────────────────────
#  INSTITUTE HEADER (always visible)
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="inst-banner">
  <div class="inst-name">🎓 {INSTITUTE_NAME}</div>
  <div class="inst-city">📍 {INSTITUTE_CITY}</div>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🏫 Course & Semester")

    course_options = [
        "BCA", "BBA", "MBA", "MCA", "B.Sc (CS)", "B.Sc (IT)",
        "B.Com", "M.Com", "B.Tech (CS)", "B.Tech (IT)", "Other"
    ]
    course = st.selectbox("Course", course_options)
    if course == "Other":
        course = st.text_input("Enter Course Name", placeholder="e.g. BBA (Hons)")

    sem_num = st.selectbox("Semester", list(range(1, 9)), index=4)
    sem_label = ordinal(sem_num)

    st.markdown("---")
    st.markdown("## 📂 Upload Result File")
    uploaded = st.file_uploader(
        "Result Format Excel (.xlsx)", type=['xlsx'],
        help="Sheet 1: Student marks & status | Sheet 2: Subject–Faculty mapping"
    )
    st.markdown("---")
    st.markdown("""
### 📌 How to use
1. Select **Course** and **Semester**
2. Upload the **Result Format Excel**
3. Analysis generates automatically
4. Explore the tabs for charts & tables
5. Export as **Excel** or **PDF report**
---
### 📄 File Format
**Sheet 1** – Roll No., Name, Subject marks (Ext/Int/Ttl), Total, Percentage, Status  
**Sheet 2** – Subject Code, Subject Name, Faculty Name, Sections
""")

# ─────────────────────────────────────────────────────────────────────────────
#  DYNAMIC REPORT TITLE BAR
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    f'<div class="report-bar">📊 Result Analysis System &nbsp;|&nbsp; '
    f'{course} &nbsp;|&nbsp; {sem_label} Semester</div>',
    unsafe_allow_html=True
)

# ─────────────────────────────────────────────────────────────────────────────
#  UPLOAD PROMPT
# ─────────────────────────────────────────────────────────────────────────────
if not uploaded:
    st.markdown("""
    <div class="upload-area">
      <h3 style="color:#1a237e;margin-bottom:.5rem">📤 Upload Result Format Excel</h3>
      <p style="color:#444;margin-bottom:.3rem">Use the <strong>sidebar</strong> to upload your result file</p>
      <p style="color:#777;font-size:.87rem">Sheet 1: Student Data &nbsp;|&nbsp; Sheet 2: Subject–Faculty Map</p>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
#  PROCESS FILE
# ─────────────────────────────────────────────────────────────────────────────
with st.spinner("⏳ Processing your Excel file…"):
    try:
        df      = parse_sheet1(uploaded)
        sf_df   = parse_sheet2(uploaded)
        summary = result_summary(df, sem_label)
        sdf     = compute_subject_stats(df, sf_df)
        fdf     = compute_faculty_stats(df, sf_df)
        gdf     = compute_grade_dist(df, sf_df)
        top_df  = compute_toppers(df, 10)
        subj_detected = detect_subject_ttl_cols(df)
    except Exception as e:
        st.error(f"❌ Error processing file: {e}")
        st.exception(e)
        st.stop()

st.success(
    f"✅ File processed — **{len(df)} students**, "
    f"**{len(subj_detected)} subjects detected** ({', '.join(subj_detected.keys())})"
)

# ─────────────────────────────────────────────────────────────────────────────
#  KPI ROW
# ─────────────────────────────────────────────────────────────────────────────
pass_key  = f'Pass ({sem_label})'
pct_key   = f'% Pass in {sem_label}'
kpi_defs = [
    ('Total Students',      summary.get('Total Registered', 0),          '#1a237e'),
    (pass_key,              summary.get(pass_key, 0),                    '#2e7d32'),
    ('Later / Overall',     summary.get('Later / Overall Pass', 0),      '#1565c0'),
    ('Back Paper',          summary.get('Back Paper', 0),                '#e65100'),
    ('Fail',                summary.get('Fail', 0),                      '#c62828'),
    (f'Pass % ({sem_label})', f"{summary.get(pct_key, 0)}%",            '#6a1b9a'),
]
cols = st.columns(6)
for col, (lbl, val, color) in zip(cols, kpi_defs):
    with col:
        st.markdown(
            f'<div class="kpi" style="border-color:{color}">'
            f'<div class="kpi-val" style="color:{color}">{val}</div>'
            f'<div class="kpi-lbl">{lbl}</div></div>',
            unsafe_allow_html=True
        )

st.markdown("---")

# ─────────────────────────────────────────────────────────────────────────────
#  TABS
# ─────────────────────────────────────────────────────────────────────────────
t1, t2, t3, t4, t5, t6 = st.tabs([
    "📊 Overview",
    "📚 Subject Analysis",
    "👩‍🏫 Faculty Analysis",
    "📈 Grade Distribution",
    "🏆 Toppers",
    "📋 Student Data",
])

# ── Tab 1: Overview ───────────────────────────────────────────────────────────
with t1:
    st.markdown('<div class="sec">Overall Result Summary</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        sum_df = pd.DataFrame(list(summary.items()), columns=['Metric', 'Value']).set_index('Metric')
        st.dataframe(sum_df, use_container_width=True, height=330)
    with c2:
        if summary:
            st.plotly_chart(chart_pie(summary), use_container_width=True)
    if summary:
        st.plotly_chart(chart_bar_summary(summary), use_container_width=True)
    h = chart_score_hist(df)
    if h:
        st.plotly_chart(h, use_container_width=True)

# ── Tab 2: Subject Analysis ───────────────────────────────────────────────────
with t2:
    st.markdown('<div class="sec">Subject Wise Result Analysis</div>', unsafe_allow_html=True)
    if not sdf.empty:
        st.dataframe(sdf, use_container_width=True, hide_index=True)
        c1, c2 = st.columns(2)
        with c1: st.plotly_chart(chart_subject_pass(sdf), use_container_width=True)
        with c2: st.plotly_chart(chart_avg(sdf), use_container_width=True)
    else:
        st.info("No subject data detected. Ensure Sheet 1 has columns like '501_Ttl', '502_Ttl', etc.")

# ── Tab 3: Faculty Analysis ───────────────────────────────────────────────────
with t3:
    st.markdown('<div class="sec">Faculty Wise Result Analysis</div>', unsafe_allow_html=True)
    if not fdf.empty:
        st.dataframe(fdf, use_container_width=True, hide_index=True)
        st.plotly_chart(chart_faculty(fdf), use_container_width=True)
    else:
        st.info("No faculty data found. Ensure Sheet 2 has Subject Code, Faculty Name, and Sections columns.")

# ── Tab 4: Grade Distribution ─────────────────────────────────────────────────
with t4:
    st.markdown('<div class="sec">Score Band Distribution by Subject</div>', unsafe_allow_html=True)
    if not gdf.empty:
        disp = ['Subject','< 50','50–59.9','60–74.9','≥ 75','< 50 %','50–59.9 %','60–74.9 %','≥ 75 %']
        st.dataframe(gdf[disp], use_container_width=True, hide_index=True)
        st.plotly_chart(chart_grade_dist(gdf), use_container_width=True)
    else:
        st.info("No grade distribution data.")

# ── Tab 5: Toppers ────────────────────────────────────────────────────────────
with t5:
    st.markdown('<div class="sec">🏆 Top Students</div>', unsafe_allow_html=True)
    if not top_df.empty:
        st.dataframe(top_df, use_container_width=True, hide_index=True)
        ct = chart_toppers(top_df)
        if ct:
            st.plotly_chart(ct, use_container_width=True)
    else:
        st.info("Could not compute toppers. Ensure a Grand Total or Total column exists in Sheet 1.")

# ── Tab 6: Student Data ───────────────────────────────────────────────────────
with t6:
    st.markdown('<div class="sec">Complete Student Result Data</div>', unsafe_allow_html=True)
    c1, c2 = st.columns([3, 1])
    with c1:
        search = st.text_input("🔍 Search by name, roll no., status, or group", "")
    with c2:
        status_col_name = find_col(df, ['status'])
        if status_col_name:
            all_statuses = ['All'] + sorted(df[status_col_name].dropna().astype(str).str.upper().unique().tolist())
            status_filter = st.selectbox("Filter by Status", all_statuses)
        else:
            status_filter = 'All'

    disp_df = df.copy()
    if search:
        mask = disp_df.astype(str).apply(
            lambda x: x.str.contains(search, case=False, na=False)
        ).any(axis=1)
        disp_df = disp_df[mask]
    if status_filter != 'All' and status_col_name:
        disp_df = disp_df[disp_df[status_col_name].astype(str).str.upper() == status_filter]

    st.dataframe(disp_df, use_container_width=True, hide_index=True, height=480)
    st.caption(f"Showing {len(disp_df)} of {len(df)} students")

# ─────────────────────────────────────────────────────────────────────────────
#  EXPORT OPTIONS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("### 📥 Export Options")

safe_course = re.sub(r'[^\w\s-]', '', course).replace(' ', '_')
file_stem   = f"Result_Analysis_{safe_course}_{sem_label}_Sem"

d1, d2 = st.columns(2)
with d1:
    excel_data = build_excel(course, sem_label, summary, sdf, fdf, gdf, top_df, df)
    st.download_button(
        label="⬇️ Download Excel Analysis (.xlsx)",
        data=excel_data,
        file_name=f"{file_stem}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with d2:
    html_data = build_print_html(course, sem_label, summary, sdf, fdf, gdf, top_df)
    st.download_button(
        label="🖨️ Download Print-Ready PDF Report (.html)",
        data=html_data,
        file_name=f"{file_stem}_Report.html",
        mime="text/html",
        use_container_width=True,
        help="Open in browser → Ctrl+P → Save as PDF (select Black & White if needed)",
    )

st.markdown("""
<div class="info-box">
  <strong>🖨️ How to save as PDF:</strong>
  Download the HTML report → open in any browser (Chrome / Edge / Firefox)
  → press <strong>Ctrl+P</strong> (Windows) or <strong>⌘P</strong> (Mac)
  → select <em>"Save as PDF"</em> as the printer → choose <em>Black &amp; White</em> if needed → Click Print.
</div>
""", unsafe_allow_html=True)
